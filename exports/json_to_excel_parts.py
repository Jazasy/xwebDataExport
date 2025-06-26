# json_to_excel_parts.py
# Script 2: merged-part-xx.json fájlokat egyenként külön Excel fájlokká alakít

import json
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

# Minden 'merged-part-*.json' fájl feldolgozása
part_files = sorted(glob.glob("merged-part-*.json"))

for part_file in part_files:
    print(f"📦 Feldolgozás: {part_file}")
    with open(part_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    device_sheet = wb.active
    device_sheet.title = "Devices"
    device_sheet.append(["Device ID", "Device Name", "Link to Points"])

    for device_index, device in enumerate(data):
        device_id = device.get("id", "")
        device_name = device.get("name", "")

        points_sheet_name = f"Dev{device_index+1}_Points"
        ps = wb.create_sheet(points_sheet_name[:31])
        ps.append(["Point ID", "Label", "Unit", "Link to Values"])

        for point_index, point in enumerate(device.get("points", [])):
            point_id = point.get("id", "")
            label = point.get("label", "")
            unit = point.get("unit", "")
            values = point.get("value", [])

            value_sheet_name = f"Dev{device_index+1}_Pt{point_index+1}_Values"
            vs = wb.create_sheet(value_sheet_name[:31])
            vs.append(["Timestamp", "DateTime", "Value"])

            for entry in values:
                ts = entry.get("timestamp", "")
                dt = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, (int, float)) else ""
                val = entry.get("value", "")
                vs.append([ts, dt, val])

            ps.append([point_id, label, unit, "Show"])
            cell = ps.cell(row=ps.max_row, column=4)
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{value_sheet_name[:31]}'!A1", display="Show")
            cell.style = "Hyperlink"

        device_sheet.append([device_id, device_name, "Show"])
        cell = device_sheet.cell(row=device_sheet.max_row, column=3)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{points_sheet_name[:31]}'!A1", display="Show")
        cell.style = "Hyperlink"

    out_file = part_file.replace(".json", ".xlsx")
    wb.save(out_file)
    print(f" Excel mentve: {out_file}")
