import json
import os
from datetime import datetime
from openpyxl import Workbook
from tqdm import tqdm

def timestamp_to_str(ts):
    return datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')

# Beolvasás
with open("final-merged.json", "r", encoding="utf-8") as f:
    data = json.load(f)

wb = Workbook()
ws_devices = wb.active
ws_devices.title = "Devices"
ws_devices.append(["device_id", "device_name"])

ws_points = wb.create_sheet("Points")
ws_points.append(["point_id", "device_id", "label", "unit"])

print("💾 Devices és Points sheet készítése...")
for device in tqdm(data):
    device_id = device.get("id")
    device_name = device.get("name")
    ws_devices.append([device_id, device_name])

    for point in device.get("points", []):
        point_id = point.get("id")
        label = point.get("label", "")
        unit = point.get("unit", "")
        ws_points.append([point_id, device_id, label, unit])

        # Sheet a value-knak
        sheet_name = f"dev{device_id}_pt{point_id}_val"
        # Excel limit workaround: max 31 karakter egy sheet neve
        sheet_name = sheet_name[:31]

        ws_val = wb.create_sheet(title=sheet_name)
        ws_val.append(["timestamp", "date", "value"])

        for entry in point.get("value", []):
            ts = entry.get("timestamp")
            val = entry.get("value")
            date_str = timestamp_to_str(ts)
            ws_val.append([ts, date_str, val])

# Mentés
output_name = "exported_data.xlsx"
wb.save(output_name)
print(f"Kész! Excel mentve: {output_name}")
